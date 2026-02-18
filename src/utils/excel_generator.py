"""
Excel Generator Module for Agentic AI Customer Support
FIXED: Auto-creates exports folder and generates Excel files
"""
import os
from datetime import datetime
from typing import List, Dict
import sys
import csv

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from src.database.database import db_manager

# Try to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will use CSV fallback.")


class ExcelGenerator:
    """Generates Excel reports for customer support data"""
    
    def __init__(self):
        # FIXED: Create exports folder in project root (same level as src/)
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        self.export_path = os.path.join(project_root, "exports")
        
        # Ensure exports directory exists
        self._ensure_export_directory()
        
        print(f"âœ… Excel exports will be saved to: {self.export_path}")
    
    def _ensure_export_directory(self):
        """Create exports directory if it doesn't exist"""
        try:
            if not os.path.exists(self.export_path):
                os.makedirs(self.export_path, exist_ok=True)
                print(f"âœ… Created exports directory: {self.export_path}")
            else:
                print(f"âœ… Exports directory exists: {self.export_path}")
        except Exception as e:
            print(f"âŒ Error creating exports directory: {e}")
            # Fallback to current directory
            self.export_path = os.path.join(os.getcwd(), "exports")
            os.makedirs(self.export_path, exist_ok=True)
            print(f"âš ï¸  Using fallback exports directory: {self.export_path}")
    
    def generate_chat_summaries_report(self, summaries: List[Dict] = None) -> str:
        """Generate Excel report of chat summaries"""
        # Ensure directory exists before generating
        self._ensure_export_directory()
        
        if summaries is None:
            summaries = db_manager.get_all_chat_summaries()
        
        if not summaries:
            print("âš ï¸  No chat summaries found - creating empty report")
            return self._create_empty_report("chat_summaries", "No chat summaries found")
        
        headers = [
            "Session ID", "User ID", "Issue Type", "Issue Description",
            "Proposed Solution", "Resolution Status", "Customer Sentiment",
            "Action Items", "Message Count", "Shipment ID", "Shipment Type", "Timestamp"
        ]
        
        rows = []
        for summary in summaries:
            rows.append([
                summary.get("session_id", ""),
                summary.get("user_id", ""),
                summary.get("issue_type", ""),
                summary.get("issue_description", ""),
                summary.get("proposed_solution", ""),
                summary.get("resolution_status", ""),
                summary.get("customer_sentiment", ""),
                ", ".join(summary.get("action_items", [])) if isinstance(summary.get("action_items"), list) else str(summary.get("action_items", "")),
                summary.get("message_count", 0),
                summary.get("shipment_id", "N/A"),
                summary.get("shipment_type", "N/A"),
                str(summary.get("timestamp", ""))
            ])
        
        filename = f"chat_summaries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_path, filename)
        
        try:
            if OPENPYXL_AVAILABLE:
                self._create_excel_file(filepath, "Chat Summaries", headers, rows)
                print(f"âœ… Chat summaries Excel created: {filepath}")
            else:
                filepath = filepath.replace('.xlsx', '.csv')
                self._create_csv_file(filepath, headers, rows)
                print(f"âœ… Chat summaries CSV created: {filepath}")
        except Exception as e:
            print(f"âŒ Error creating chat summaries report: {e}")
            import traceback
            traceback.print_exc()
        
        return filepath
    
    def generate_shipments_report(self, shipments: List[Dict] = None) -> str:
        """Generate Excel report of shipments"""
        # Ensure directory exists before generating
        self._ensure_export_directory()
        
        if shipments is None:
            shipments = db_manager.get_all_shipments()
        
        if not shipments:
            print("âš ï¸  No shipments found - creating empty report")
            return self._create_empty_report("shipments", "No shipments found")
        
        headers = [
            "Shipment ID", "Type", "User ID", "Order ID", "Product ID",
            "Status", "Address", "Created At", "Estimated Completion", "Current Stage"
        ]
        
        rows = []
        for shipment in shipments:
            rows.append([
                shipment.get("shipment_id", ""),
                shipment.get("type", "").title() if shipment.get("type") else "",
                shipment.get("user_id", ""),
                shipment.get("order_id", ""),
                shipment.get("product_id", ""),
                shipment.get("status", ""),
                shipment.get("address", ""),
                str(shipment.get("created_at", "")),
                str(shipment.get("estimated_completion", "")),
                shipment.get("current_stage", {}).get("name", "") if isinstance(shipment.get("current_stage"), dict) else ""
            ])
        
        filename = f"shipments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_path, filename)
        
        try:
            if OPENPYXL_AVAILABLE:
                self._create_excel_file(filepath, "Shipments", headers, rows)
                print(f"âœ… Shipments Excel created: {filepath}")
            else:
                filepath = filepath.replace('.xlsx', '.csv')
                self._create_csv_file(filepath, headers, rows)
                print(f"âœ… Shipments CSV created: {filepath}")
        except Exception as e:
            print(f"âŒ Error creating shipments report: {e}")
            import traceback
            traceback.print_exc()
        
        return filepath
    
    def _create_excel_file(self, filepath: str, sheet_name: str, 
                           headers: List[str], rows: List[List]) -> None:
        """Create an Excel file with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            max_length = len(str(headers[col - 1]))
            for row in range(2, len(rows) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Save file
        wb.save(filepath)
        print(f"âœ… Excel file saved: {filepath}")
    
    def _create_csv_file(self, filepath: str, headers: List[str], rows: List[List]) -> None:
        """Create a CSV file as fallback"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        print(f"âœ… CSV file saved: {filepath}")
    
    def _create_empty_report(self, report_type: str, message: str) -> str:
        """Create an empty report with a message"""
        filename = f"{report_type}_empty_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        filepath = os.path.join(self.export_path, filename)
        
        try:
            with open(filepath, 'w') as f:
                f.write(f"{message}\n")
                f.write(f"Generated at: {datetime.now().isoformat()}\n")
            print(f"âœ… Empty report created: {filepath}")
        except Exception as e:
            print(f"âŒ Error creating empty report: {e}")
        
        return filepath
    
    def generate_all_reports(self) -> Dict[str, str]:
        """Generate all reports and return file paths"""
        print("\n" + "="*60)
        print("  GENERATING EXCEL REPORTS")
        print("="*60)
        
        reports = {
            "chat_summaries": self.generate_chat_summaries_report(),
            "shipments": self.generate_shipments_report()
        }
        
        print("\nâœ… All reports generated successfully!")
        print(f"ğŸ“ Reports saved to: {self.export_path}")
        print("="*60 + "\n")
        
        return reports


# Create singleton instance
excel_generator = ExcelGenerator()